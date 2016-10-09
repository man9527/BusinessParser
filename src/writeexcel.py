import collections
import os
from openpyxl import Workbook
from openpyxl.styles import colors, PatternFill
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side

class CompanyDataWriter:
    def __init__(self, fileName):
        self.headers = collections.OrderedDict({})
        self.groupHeaders = collections.OrderedDict({})

        self.headers["ID"]="ID"
        self.headers["company name"]="Company Name"
        self.headers["YEAR"]="Year"
        self.headers["class"]="Class"
        self.headers["company type"]="Company Type"
        self.headers["parent name"] = "Parent Name"
        self.headers["ultimate parent name"] = "Ultimate Parent Name"
        self.headers["ult. parent hierarchy"] = "Ult. Parent Hierarchy"
        self.headers["street"] = "Street"
        self.headers["city"] = "City"
        self.headers["state"] = "State"
        self.headers["zip"] = "Zip"
        self.headers["province"] = "Province"
        self.headers["country"] = "Country"
        self.headers["state incorporated"] = "State Incorporated"
        self.headers["percent owned by parent"] = "Percent Owned By Parent"
        self.headers["fiscal year end"] = "Fiscal Year End"
        self.headers["sales"] = "Sales"
        self.headers["assets"] = "Assets"
        self.headers["liabilities"] = "Liabilities"
        self.headers["net worth"] = "Net Worth"
        self.headers["net income"] = "Net Income"
        self.headers["number of employees"] = "Number Of Employees"
        self.headers["year founded"] = "Year Founded"
        self.groupHeaders["trade name"] = "Trade Name"
        self.groupHeaders["sic"] = "Sic"
        self.groupHeaders["stock exchange"] = "Stock Exchange"
        self.groupHeaders["ticker symbol"] = "Ticker Symbol"
        self.groupHeaders["outside firm"] = "Outside Firm"
        self.groupHeaders["function outside firm"] = "Outside Firm Function"
        self.groupHeaders["executive"] = "Executive"
        self.groupHeaders["function executive"] = "Executive Function"
        self.groupHeaders["director"] = "Director"
        self.groupHeaders["function director"] = "Director Function"

        self.fileName = fileName
        self.wb = Workbook()
        self.ws = self.wb.active

        self.columnCount = 2
        self.rowCount = 1

        # populate headers
        headers = []

        for key, value in self.headers.items():
            headers.append(value)
            self.columnCount += 1

        for key, value in self.groupHeaders.items():
            headers.append(value)
            self.columnCount += 1

        self.ws.append(headers)

    def append(self, company, result):

        # populate data
        for year, data in result.items():
            row=[]
            row.append(company.id)
            row.append(data["company name"])
            row.append(year)

            for key, value in self.headers.items():
                if key != "ID" and key != "company name" and key != "YEAR":
                    if key in data.keys():
                        row.append(data[key])
                    else:
                        row.append("")

            for key, value in self.groupHeaders.items():
                row.append(self.__handleGroupValue__(key, data))

            self.ws.append(row)
            self.rowCount+=1

    def __handleGroupValue__(self, groupKey, data):

        result = ""

        for key, value in data.items():
            if (key.startswith(groupKey) and value):
                result = result + data[key] + "\n"

        return result

    def append2(self, company, result):
        # populate headers
        data = result.itervalues().next()
        headers=["ID","YEAR"]
        for value in data:
            headers.append(str(value).upper())
            self.columnCount+=1

        self.ws.append(headers)

        # populate data
        for year, data in result.items():
            row=[]
            row.append(company.id)
            row.append(year)
            for key, value in data.items():
                row.append(value)
            self.ws.append(row)
            self.rowCount+=1

    def save(self):
        # add style
        redFill = PatternFill(start_color='cdd0d1',
                              end_color='cdd0d1',
                              fill_type='solid')

        ft_red = Font(color=colors.RED)

        for i in range(1,self.columnCount+1):
            self.ws.cell(row=1, column=i).fill = redFill;

        for i in range(1,self.rowCount+1):
            self.ws.cell(row=i, column=1).font = ft_red;

        for col in ["X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG"]:
            for i in range(1,self.rowCount+1):
                self.ws.cell(col+str(i)).alignment = Alignment(wrapText=True, vertical="top");

        for col in ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W"]:
            for i in range(1,self.rowCount+1):
                self.ws.cell(col+str(i)).alignment = Alignment(vertical="top");

        dims = {}
        for row in self.ws.rows:
            for cell in row:
                if cell.value:
                    values = cell.value.split("\n")
                    maxLen = 0
                    for value in values:
                        maxLen=max(maxLen, len(value))

                    dims[cell.column] = max((dims.get(cell.column, 0), maxLen+5))
        for col, value in dims.items():
            self.ws.column_dimensions[col].width = value

        self.wb.save(self.fileName)


class HierarchyWriter:
    def write(self, company, result):
        wb = Workbook()
        ws = wb.active

        rowCount = 1

        # populate headers
        headers = ["Year", "Company Name", "ID", "City", "State", "Level", "Role",
                   "Upper Comany Name", "Upper Company ID", "Upper Company City", "Upper Company State", "Upper Company Level", "Upper Company Role",
                   "Ultimate Comany Name", "Ultimate Company ID", "Ultimate Company City", "Ultimate Company State",
                   "Ultimate Company Level", "Ultimate Company Role"]

        ws.append(headers)

        ultimate_company_name=""

        # populate company
        for key,nodes in result.items():
            for node in nodes:
                rowCount+=1
                data = [key]
                data.append(node.name)
                data.append(node.id)
                data.append(node.city)
                data.append(node.state)
                data.append(node.levelCount/2)
                data.append(node.type)

                parent = node.parent
                data.append(parent.name)
                data.append(parent.id)
                data.append(parent.city)
                data.append(parent.state)
                data.append(parent.levelCount/2)
                data.append(parent.type)

                ultimate = nodes[0]
                data.append(ultimate.name)
                ultimate_company_name = ultimate.name
                data.append(ultimate.id)
                data.append(ultimate.city)
                data.append(ultimate.state)
                data.append(ultimate.levelCount/2)
                data.append(ultimate.type)

                ws.append(data)

        # add style
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        firstColor = PatternFill(start_color='B0E5F7',
                                  end_color='B0E5F7',
                                 fill_type='solid')

        currentColor = PatternFill(start_color='BADB9C',
                            end_color='BADB9C',
                            fill_type='solid')

        parentColor = PatternFill(start_color='EDDDD3',
                            end_color='EDDDD3',
                            fill_type='solid')

        ultimateColor = PatternFill(start_color='F7C388',
                            end_color='F7C388',
                            fill_type='solid')

        for i in range(1, rowCount+1):
            ws.cell(row=i, column=1).fill = firstColor;
            ws.cell(row=i, column=1).border=thin_border

            for j in range(2, 8):
                ws.cell(row=i, column=j).fill = currentColor;
                ws.cell(row=i, column=j).border = thin_border

            for j in range(8, 14):
                ws.cell(row=i, column=j).fill = parentColor;
                ws.cell(row=i, column=j).border = thin_border

            for j in range(14, 20):
                ws.cell(row=i, column=j).fill = ultimateColor;
                ws.cell(row=i, column=j).border = thin_border

        dir_path = os.path.dirname(os.path.realpath(company.exportFileName))

        wb.save( os.path.join(dir_path, ultimate_company_name + ".xlsx"))
